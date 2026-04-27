"""
Reconciliation Agent - AWS Strands + Bedrock + FastAPI
======================================================
Exposes a FastAPI server that lets calling clients trigger a
reconciliation agent powered by AWS Bedrock (Claude) via the
Strands Agents SDK.
 
The agent reconciles the 'Timesheet Hrs' column in the Excel burnsheet
based on empId or name specified in the user prompt.
 
Environment variables required (loaded from .env):
  - AWS_BEARER_TOKEN_BEDROCK : pre-signed bearer token for Bedrock
  - AWS_SECRET_KEY           : (optional) KMS ARN – not used directly
  - AWS_REGION               : AWS region  (default: us-west-1)
 
Start:
    python Reconcillation.py
"""
 
import os, sys, json, base64, traceback, re
import importlib
from typing import Optional, List, Dict
from urllib.parse import parse_qs, urlparse, unquote
from datetime import datetime, timedelta, timezone
from pathlib import Path
 
# ---------------------------------------------------------------------------
# 1. Environment & credential bootstrap
# ---------------------------------------------------------------------------
try:
    _dotenv = importlib.import_module("dotenv")
    load_dotenv = getattr(_dotenv, "load_dotenv", lambda: None)
    load_dotenv()
except Exception:
    # dotenv isn't required at runtime — allow operation without it.
    def load_dotenv():
        return None
 
BEARER_TOKEN = os.getenv("AWS_BEARER_TOKEN_BEDROCK", "")
REGION = os.getenv("AWS_REGION", "us-west-1")
 
EXCEL_PATH = Path(__file__).resolve().parent / "public" / "Combined-Input.xlsx"
 
REQUIRED_ENV_VARS = ["AWS_BEARER_TOKEN_BEDROCK", "AWS_REGION"]
 
def _extract_credentials_from_bearer_token():
    """Decode the pre-signed-URL bearer token and set temporary AWS creds."""
    encoded = BEARER_TOKEN
    if encoded.startswith("bedrock-api-key-"):
        encoded = encoded[len("bedrock-api-key-"):]
    padding = 4 - len(encoded) % 4
    if padding != 4:
        encoded += "=" * padding
    decoded_url = base64.b64decode(encoded).decode()
    if not decoded_url.startswith("http"):
        decoded_url = "https://" + decoded_url
    parsed = urlparse(decoded_url)
    params = parse_qs(parsed.query)
    credential = params.get("X-Amz-Credential", [""])[0]
    cred_parts = credential.split("/")
    access_key = cred_parts[0]
    cred_region = cred_parts[2]
    security_token = unquote(params.get("X-Amz-Security-Token", [""])[0])
    amz_date = params.get("X-Amz-Date", [""])[0]
    expires = int(params.get("X-Amz-Expires", ["0"])[0])
 
    # Check expiry
    token_time = datetime.strptime(amz_date, "%Y%m%dT%H%M%SZ").replace(tzinfo=timezone.utc)
    expiry_time = token_time + timedelta(seconds=expires)
    now = datetime.now(timezone.utc)
    if now > expiry_time:
        raise RuntimeError(f"Bearer token expired at {expiry_time}  (now={now})")
 
    # Inject temporary creds so boto3 / strands can pick them up
    os.environ["AWS_ACCESS_KEY_ID"] = access_key
    os.environ["AWS_SECRET_ACCESS_KEY"] = access_key  # placeholder – signing done via token
    os.environ["AWS_SESSION_TOKEN"] = security_token
    os.environ["AWS_DEFAULT_REGION"] = cred_region
    return {
        "access_key": access_key,
        "region": cred_region,
        "expires": str(expiry_time),
        "remaining": str(expiry_time - now),
    }
 
 
# ---------------------------------------------------------------------------
# 2. Strands tool -- reconcile timesheet in Excel
# ---------------------------------------------------------------------------
import openpyxl
try:
    from strands import Agent, tool  # type: ignore
    from strands.models.bedrock import BedrockModel  # type: ignore
    _STRANDS_AVAILABLE = True
except Exception:
    # Provide safe fallbacks so runtime usage that doesn't need strands can continue.
    _STRANDS_AVAILABLE = False
 
    Agent = None
 
    def tool(fn=None, **kwargs):
        # If used as a decorator, return the original function so direct calls still work.
        return fn
 
    class BedrockModel:
        def __init__(self, *args, **kwargs):
            raise RuntimeError(
                "BedrockModel is unavailable: install 'strands-agents' and 'strands-agents-bedrock'"
            )
 
 
@tool
def reconcile_timesheet(
    identifier: str,
    new_timesheet_hrs: float,
    file_path: str = str(EXCEL_PATH),
) -> str:
    """Reconcile the 'Timesheet Hrs' for a specific employee in the Excel burnsheet.
 
    Looks up the employee by empId or name (case-insensitive partial match),
    updates the Timesheet Hrs column, recalculates Actual Rate, Variance,
    and monthly columns, saves the workbook, and returns a JSON summary
    including the updated row data so the UI can refresh.
 
    Args:
        identifier: Employee ID or name to search for.
        new_timesheet_hrs: The new timesheet hours value to set.
        file_path: Path to the Excel file (defaults to public/Combined-Input.xlsx).
    """
    fp = Path(file_path)
    if not fp.exists():
        return json.dumps({"error": f"File not found - {fp}"})
 
    wb = openpyxl.load_workbook(fp)
    ws = wb.active
 
    # Build header map
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
 
    # Required columns
    required = ["EMPId", "Name", "Timesheet", "Hourly Rate($)", "Projected Rate($)",
                "Actual Rate", "Variance"]
    missing_cols = [c for c in required if c not in col_map]
    if missing_cols:
        return json.dumps({"error": f"Missing columns: {missing_cols}. Available: {list(col_map.keys())}"})
 
    # Search for matching row(s)
    id_col = col_map["EMPId"]
    name_col = col_map["Name"]
    search = identifier.strip().lower()
    matched_rows = []
 
    for row_num in range(2, ws.max_row + 1):
        raw_id = str(ws.cell(row=row_num, column=id_col).value or "").strip()
        emp_id = raw_id[:-2] if raw_id.endswith(".0") else raw_id
        emp_name = str(ws.cell(row=row_num, column=name_col).value or "").strip()
        if search == emp_id.lower() or search in emp_name.lower():
            matched_rows.append(row_num)
 
    if not matched_rows:
        return json.dumps({"error": f"No employee found matching '{identifier}'."})
 
    ts_col = col_map["Timesheet"]
    rate_usd_col = col_map["Hourly Rate($)"]
    proj_col = col_map["Projected Rate($)"]
    act_col = col_map["Actual Rate"]
    var_col = col_map["Variance"]
 
    updates = []
    for row_num in matched_rows:
        old_hrs = ws.cell(row=row_num, column=ts_col).value or 0
        rate_usd = float(ws.cell(row=row_num, column=rate_usd_col).value or 0)
        projected = float(ws.cell(row=row_num, column=proj_col).value or 0)
 
        # Update timesheet hours
        ws.cell(row=row_num, column=ts_col).value = new_timesheet_hrs
 
        # Recalculate actual rate and variance (monthly columns remain unchanged)
        new_actual = round(rate_usd * new_timesheet_hrs, 2)
        new_variance = round(new_actual - projected, 2)
 
        ws.cell(row=row_num, column=act_col).value = new_actual
        ws.cell(row=row_num, column=var_col).value = new_variance
 
        emp_id_raw = str(ws.cell(row=row_num, column=id_col).value or "").strip()
        # Remove trailing .0 from float-like IDs (e.g. "2298348.0" -> "2298348")
        if emp_id_raw.endswith(".0"):
            emp_id_raw = emp_id_raw[:-2]
        emp_id = emp_id_raw
        emp_name = str(ws.cell(row=row_num, column=name_col).value or "").strip()
 
        updates.append({
            "row": row_num,
            "empId": emp_id,
            "name": emp_name,
            "oldTimesheetHrs": float(old_hrs),
            "newTimesheetHrs": new_timesheet_hrs,
            "rateUsd": rate_usd,
            "projectedRate": projected,
            "newActualRate": new_actual,
            "newVariance": new_variance,
        })
 
    wb.save(fp)
    wb.close()
 
    return json.dumps({
        "status": "success",
        "message": f"Reconciled {len(updates)} row(s).",
        "updates": updates,
    })
 
 
@tool
def read_employee_data(
    identifier: str = "",
    file_path: str = str(EXCEL_PATH),
) -> str:
    """Read employee data from the Excel burnsheet.
 
    If identifier is provided, returns matching rows. Otherwise returns
    a summary of all data.
 
    Args:
        identifier: Optional employee ID or name to search for.
        file_path: Path to the Excel file.
    """
    fp = Path(file_path)
    if not fp.exists():
        return json.dumps({"error": f"File not found - {fp}"})
 
    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
 
    if not identifier.strip():
        return json.dumps({"total_rows": ws.max_row - 1, "columns": list(col_map.keys())})
 
    id_idx = col_map.get("EMPId")
    name_idx = col_map.get("Name")
    if id_idx is None or name_idx is None:
        return json.dumps({"error": "Missing EMPId or Name column."})
 
    search = identifier.strip().lower()
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_id = str(row[id_idx] or "").strip().lower()
        emp_name = str(row[name_idx] or "").strip().lower()
        if search == emp_id or search in emp_name:
            results.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
 
    wb.close()
    return json.dumps({"matches": len(results), "data": results[:10]}, default=str)
 
 
# ---------------------------------------------------------------------------
# 3. Build the Strands Agent (lazy -- created on first use)
# ---------------------------------------------------------------------------
_agent = None
 
def _get_agent():
    global _agent
    if _agent is None:
        creds = _extract_credentials_from_bearer_token()
        model = BedrockModel(
            model_id="us.anthropic.claude-sonnet-4-20250514-v1:0",
            region_name=creds["region"],
        )
        _agent = Agent(
            model=model,
            tools=[reconcile_timesheet, read_employee_data],
            system_prompt=(
                "You are a Reconciliation Agent for a timesheet burnsheet system.\n"
                "The Excel file has columns: ESA ID, ESA Description, Verizon TQ ID, "
                "Verizon TQ Description, POC, EMPId, Name, Location, Country, ACT/PCT, "
                "Skill Set, Verizon Level Mapping, Classification, Key, Cognizant Designation, "
                "Service Line, Timesheet, Hourly Rate(Rs), Hourly Rate($), Projected Rate($), "
                "Actual Rate, Variance.\n\n"
                "When asked to reconcile, use the 'reconcile_timesheet' tool to update "
                "the Timesheet Hrs for the specified employee. Extract the employee "
                "identifier (empId or name) and the new hours value from the user prompt.\n"
                "If the user asks to look up data, use 'read_employee_data'.\n"
                "Always report the old and new values clearly in your response.\n"
                "Return your final answer as a structured summary."
            ),
        )
    return _agent
 
 
# ---------------------------------------------------------------------------
# 4. FastAPI application
# ---------------------------------------------------------------------------
try:
    _fastapi = importlib.import_module("fastapi")
    FastAPI = getattr(_fastapi, "FastAPI")
    HTTPException = getattr(_fastapi, "HTTPException")
    _fastapi_cors = importlib.import_module("fastapi.middleware.cors")
    CORSMiddleware = getattr(_fastapi_cors, "CORSMiddleware")
    _pydantic = importlib.import_module("pydantic")
    BaseModel = getattr(_pydantic, "BaseModel")
    _FASTAPI_AVAILABLE = True
except Exception:
    # FastAPI / Pydantic aren't available. Provide minimal placeholders so
    # importing this module won't crash tooling — endpoints will raise if used.
    _FASTAPI_AVAILABLE = False
 
    class HTTPException(Exception):
        def __init__(self, status_code: int = 500, detail: str = ""):
            super().__init__(detail)
            self.status_code = status_code
            self.detail = detail
 
    class BaseModel:
        def __init__(self, **kwargs):
            for k, v in kwargs.items():
                setattr(self, k, v)
 
    class FastAPI:  # minimal stub
        def __init__(self, *args, **kwargs):
            pass
 
        def add_middleware(self, *args, **kwargs):
            pass
 
    class CORSMiddleware:  # stub
        pass
 
app = FastAPI(title="Reconciliation Agent API", version="1.0.0")
try:
    app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
except Exception:
    # If middleware isn't available (stubbed), ignore.
    pass
 
 
class AgentRequest(BaseModel):
    prompt: str = "Reconcile timesheet hours for employee 2298348 to 160 hours."
 
 
class AgentResponse(BaseModel):
    status: str
    response: str
    updates: Optional[List[dict]] = None
 
 
class DirectReconcileRequest(BaseModel):
    identifier: str
    newTimesheetHrs: float
 
 
# --- Health / connection test ---
@app.get("/health")
def health_check():
    """Test connection to AWS Bedrock. Reports missing env vars on failure."""
    missing = [v for v in REQUIRED_ENV_VARS if not os.getenv(v)]
    if missing:
        return {
            "status": "error",
            "message": f"Missing environment variables: {missing}",
            "required_vars": REQUIRED_ENV_VARS,
        }
    try:
        creds = _extract_credentials_from_bearer_token()
        return {"status": "ok", "credentials": creds}
    except Exception as e:
        return {"status": "error", "message": str(e), "required_vars": REQUIRED_ENV_VARS}
 
 
# --- Run the agent (LLM-powered) ---
@app.post("/reconcile", response_model=AgentResponse)
def run_reconciliation(req: AgentRequest):
    """Send a prompt to the Strands reconciliation agent."""
    try:
        agent = _get_agent()
        result = agent(req.prompt)
 
        # Try to extract structured updates from tool output
        updates = None
        result_str = str(result)
        try:
            # The agent may have called reconcile_timesheet which returns JSON
            for line in result_str.split("\n"):
                line = line.strip()
                if line.startswith("{") and "updates" in line:
                    parsed = json.loads(line)
                    if "updates" in parsed:
                        updates = parsed["updates"]
                        break
        except Exception:
            pass
 
        return AgentResponse(status="success", response=result_str, updates=updates)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
 
 
# --- Direct (non-LLM) reconcile endpoint ---
@app.post("/reconcile-direct")
def run_reconciliation_direct(req: DirectReconcileRequest):
    """Directly reconcile without going through the LLM -- useful as a fallback."""
    try:
        result_json = reconcile_timesheet(
            identifier=req.identifier,
            new_timesheet_hrs=req.newTimesheetHrs,
            file_path=str(EXCEL_PATH),
        )
        result = json.loads(result_json)
        return result
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
 
 
# --- Read employee data endpoint ---
@app.get("/employee/{identifier}")
def get_employee(identifier: str):
    """Look up employee data by empId or name."""
    try:
        result_json = read_employee_data(identifier=identifier, file_path=str(EXCEL_PATH))
        return json.loads(result_json)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
 
 
# --- Read all data from Excel for the UI ---
@app.get("/data")
def get_all_data():
    """Read all rows from the Excel burnsheet and return as JSON arrays."""
    try:
        fp = EXCEL_PATH
        if not fp.exists():
            raise HTTPException(status_code=404, detail=f"Excel file not found: {fp}")
 
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
 
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(list(row))
 
        wb.close()
        return {"data": rows}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
 
 
# ---------------------------------------------------------------------------
# 5. Entrypoint
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        uvicorn = importlib.import_module("uvicorn")
    except Exception:
        uvicorn = None
 
    HOST = "0.0.0.0"
    PORT = 8000
 
    print("=" * 60)
    print("  RECONCILIATION AGENT -- Starting")
    print("=" * 60)
 
    # Pre-flight: check env vars
    missing = [v for v in REQUIRED_ENV_VARS if not os.getenv(v)]
    if missing:
        print(f"\n  [FAIL] Missing environment variables: {missing}")
        print("  Set them in your .env file and restart.\n")
        sys.exit(1)
 
    # Pre-flight: check credentials
    try:
        creds = _extract_credentials_from_bearer_token()
        print(f"  [OK] Bearer token valid - expires {creds['expires']}")
        print(f"       Region: {creds['region']}")
    except Exception as e:
        print(f"\n  [FAIL] Credential error: {e}\n")
        sys.exit(1)
 
    # Pre-flight: check Excel file
    if EXCEL_PATH.exists():
        print(f"  [OK] Excel file found: {EXCEL_PATH}")
    else:
        print(f"  [WARN] Excel file NOT found: {EXCEL_PATH}")
 
    print()
    print(f"  API URL:  http://localhost:{PORT}")
    print(f"     Health:      GET  http://localhost:{PORT}/health")
    print(f"     Reconcile:   POST http://localhost:{PORT}/reconcile")
    print(f"     Direct:      POST http://localhost:{PORT}/reconcile-direct")
    print(f"     Employee:    GET  http://localhost:{PORT}/employee/{{id_or_name}}")
    print()
    print("=" * 60)
 
    uvicorn.run(app, host=HOST, port=PORT)
 